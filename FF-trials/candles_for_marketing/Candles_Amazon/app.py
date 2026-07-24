#!/usr/bin/env python3
"""
香型分类 UI — 拖入 Excel，自动分类并展示层级统计表。
streamlit run app.py
"""

import io
import re
import sys
import time
import random
from pathlib import Path
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed

import openpyxl
import streamlit as st

sys.path.insert(0, str(Path(__file__).parent))
from scent_classifier import (classify_product, KEYWORD_TO_FAMILY, FAMILIES,
                              _classify_one_sync, cleanup_playwright)

# ============================================================
# 页面配置
# ============================================================
st.set_page_config(
    page_title="香型分类器",
    page_icon="🕯️",
    layout="wide",
)

st.title("Candles_Amazon")
st.caption("拖入 Excel → 自动搜索 Amazon → 层级统计表")

# ============================================================
# 上传文件
# ============================================================
uploaded = st.file_uploader(
    "拖入 Excel 文件（商品名在第 9 列）",
    type=["xlsx"],
    help="支持久谦中台导出的原始数据格式",
)

if uploaded is None:
    st.info("👆 请拖入 Excel 文件开始")
    st.stop()

# ============================================================
# 解析 Excel
# ============================================================
def parse_upload(uploaded_file):
    wb = openpyxl.load_workbook(uploaded_file)
    ws = wb.active
    products = []
    for row_idx in range(3, ws.max_row + 1):
        name = ws.cell(row=row_idx, column=9).value
        if name and str(name).strip():
            products.append(str(name).strip())
    return products

products = parse_upload(uploaded)
st.success(f"已读取 {len(products)} 条商品")

# ============================================================
# 处理配置
# ============================================================
col1, col2 = st.columns(2)
with col1:
    max_n = st.number_input("处理数量", min_value=1, max_value=len(products),
                            value=min(30, len(products)))
with col2:
    run_btn = st.button("▶ 开始分类", type="primary", use_container_width=True)

if not run_btn:
    st.info("设置数量后点击「开始分类」")
    st.stop()

# ============================================================
# 运行分类
# ============================================================
results = [None] * max_n
progress = st.progress(0, "正在双线程分类...")
status_text = st.empty()

SKIP_IN_SECONDARY = {"fresh", "clean", "cotton", "cake"}

def process_one(idx, product_name):
    try:
        r = _classify_one_sync(product_name)
    except Exception as e:
        r = {"产品名": product_name, "大类": f"处理失败({type(e).__name__})",
             "细分": str(e)[:100], "搜索URL": "", "方法": ""}
    # 后处理
    try:
        clean_sub = re.sub(r"\[搜\]", "", r.get("细分", ""))
        clean_sub = re.sub(r"\([^)]*\)", "", clean_sub)
        kws = [kw.strip() for kw in clean_sub.split(",") if kw.strip()]
        kws = [kw for kw in kws if kw not in SKIP_IN_SECONDARY]
        r["细分_clean"] = ", ".join(kws)
    except Exception:
        r["细分_clean"] = ""
    return idx, r

try:
    with ThreadPoolExecutor(max_workers=2) as pool:
        futures = {}
        for i in range(max_n):
            futures[pool.submit(process_one, i, products[i])] = i
            time.sleep(random.uniform(0.3, 0.6))  # 错开启动，避免同时打 Amazon
        done = 0
        for future in as_completed(futures):
            idx, r = future.result()
            results[idx] = r
            done += 1
            progress.progress(done / max_n)
            m = r.get("方法", "")
            tag = "⚡" if m == "cffi" else "🐢" if m == "pw" else ""
            status_text.text(f"[{done}/{max_n}] {tag} {r['大类'][:35]} | {products[idx][:50]}")
finally:
    cleanup_playwright()

progress.empty()
status_text.empty()
ok_count = sum(1 for r in results if r and "失败" not in r.get("大类", "")
               and "未识别" not in r.get("大类", ""))
unscented = sum(1 for r in results if r and "无香" in r.get("大类", ""))
st.success(f"✅ 完成！已分类 {ok_count} 条，无香 {unscented} 条，共 {max_n} 条")

# ============================================================
# 构建层级统计
# ============================================================
valid = [r for r in results
         if "无香" not in r["大类"]
         and "未识别" not in r["大类"]
         and "失败" not in r["大类"]]
total_valid = len(valid)

# 统计结构
# stats[fam] = {
#     "count": set(product_ids),
#     "二级": {kw: set(product_ids)}
# }
FAMILY_ORDER = list(FAMILIES.keys())
stats = {}
for fam in FAMILY_ORDER:
    stats[fam] = {"products": set(), "二级": defaultdict(lambda: {"products": set()})}

for idx, r in enumerate(valid):
    pid = idx + 1  # 1-indexed among valid products
    大类_str = r["大类"]
    细分_str = r.get("细分_clean", "")
    细分_kws = [kw.strip() for kw in 细分_str.split(",") if kw.strip()] if 细分_str else []

    # 一级
    is_mix = "混合 Mix" in 大类_str
    for fam in FAMILY_ORDER:
        if fam in 大类_str:
            stats[fam]["products"].add(pid)

    # 二级归属
    for kw in 细分_kws:
        kw_lower = kw.lower()
        if is_mix:
            # Mix 产品的细分全部归入 Mix 家族下
            stats["混合 Mix"]["二级"][kw_lower]["products"].add(pid)
        else:
            fam = KEYWORD_TO_FAMILY.get(kw_lower)
            if fam:
                stats[fam]["二级"][kw_lower]["products"].add(pid)
                stats[fam]["products"].add(pid)

# ============================================================
# 展示：层级统计表（马卡龙色系 + HTML 表格，颜色填满整格）
# ============================================================
st.subheader(f"📊 香型层级统计（有效商品 {total_valid} 个）")

MACARON = {
    "花香 Floral":    ("#FFE0E8", "#FFB8C6"),
    "清新 Fresh":     ("#D4F5E0", "#A8E6C1"),
    "木质 Woody":     ("#FFE8D0", "#FFCC99"),
    "美食 Gourmand":  ("#FFE4CC", "#FFC899"),
    "草本 Herbal":    ("#D8F0D8", "#B0E0B0"),
    "东方 Oriental":  ("#EDE0F5", "#D4B8E8"),
    "果香 Fruity":    ("#FFE0E0", "#FFB8B8"),
    "柑橘 Citrus":    ("#FFF8D0", "#FFEE99"),
    "海洋 Aquatic":   ("#D0E8FF", "#A8D4FF"),
    "混合 Mix":      ("#F0F0F0", "#D0D0D0"),
}

DEFAULT_COLORS = ("#F5F5F5", "#E0E0E0")

table_rows = []
sorted_fams = sorted(stats.items(), key=lambda x: -len(x[1]["products"]))

# ---- 构建 HTML 表格 ----
html = """<table style="border-collapse:collapse;width:100%;font-size:14px;font-family:-apple-system,BlinkMacSystemFont,sans-serif;">
<thead>
<tr style="background:#f0f0f0;text-align:left;">
<th style="padding:8px 12px;">大类</th>
<th style="padding:8px 12px;text-align:center;">数量</th>
<th style="padding:8px 12px;text-align:center;">占比</th>
<th style="padding:8px 12px;">细分</th>
<th style="padding:8px 12px;text-align:center;">数量</th>
<th style="padding:8px 12px;text-align:center;">占比</th>
<th style="padding:8px 12px;">涉及产品</th>
</tr>
</thead>
<tbody>
"""

for fam, s in sorted_fams:
    fam_count = len(s["products"])
    if fam_count == 0:
        continue
    fam_pct = round(fam_count / total_valid * 100, 1)
    colors = MACARON.get(fam, DEFAULT_COLORS)
    fam_c = colors[1] if fam_pct > 30 else colors[0]
    is_mix = (fam == "混合 Mix")

    if is_mix:
        # Mix 不展开细分，一行汇总
        pids = sorted(s["products"])
        pid_parts = []
        for pid in pids:
            r = valid[pid - 1]
            url = r.get("搜索URL", "")
            if url and url.startswith("http"):
                pid_parts.append(f'<a href="{url}" target="_blank">#{pid}</a>')
            else:
                pid_parts.append(f'<span style="color:#aaa">#{pid}</span>')
        pids_html = ", ".join(pid_parts)

        html += "<tr>"
        html += f'<td style="background:{fam_c};font-weight:600;padding:8px 12px;">{fam}</td>'
        html += f'<td style="background:{fam_c};padding:8px 12px;text-align:center;">{fam_count}</td>'
        html += f'<td style="background:{fam_c};padding:8px 12px;text-align:center;">{fam_pct}%</td>'
        html += f'<td style="background:{fam_c};padding:8px 12px;">—</td>'
        html += f'<td style="background:{fam_c};padding:8px 12px;text-align:center;">—</td>'
        html += f'<td style="background:{fam_c};padding:8px 12px;text-align:center;">—</td>'
        html += f'<td style="padding:8px 12px;">{pids_html}</td>'
        html += "</tr>\n"

        table_rows.append({
            "_fam": fam, "_fam_n": fam_count, "_fam_p": f"{fam_pct}%",
            "_kw": "", "_kw_n": "", "_kw_p": "",
            "涉及产品": re.sub(r"<[^>]+>", "", pids_html),
        })
        continue

    sorted_kw = sorted(s["二级"].items(), key=lambda x: -len(x[1]["products"]))
    active_kw = [(kw, d) for kw, d in sorted_kw if len(d["products"]) > 0]

    if not active_kw:
        active_kw = [("—", {"products": s["products"]})]

    rowspan = len(active_kw)
    first_row = True
    fallback_row = (len(active_kw) == 1 and active_kw[0][0] == "—")

    for kw, kw_data in active_kw:
        kw_is_placeholder = (kw == "—" and fallback_row)
        kw_count = len(kw_data["products"])
        kw_pct = round(kw_count / total_valid * 100, 1)
        kw_c = colors[1] if kw_pct > 15 else colors[0]

        pids = sorted(kw_data["products"])
        pid_parts = []
        for pid in pids:
            r = valid[pid - 1]
            url = r.get("搜索URL", "")
            if url and url.startswith("http"):
                pid_parts.append(f'<a href="{url}" target="_blank">#{pid}</a>')
            else:
                pid_parts.append(f'<span style="color:#aaa">#{pid}</span>')
        pids_html = ", ".join(pid_parts)

        # ---- HTML row ----
        html += "<tr>"
        if first_row:
            html += (
                f'<td style="background:{fam_c};font-weight:600;padding:8px 12px;"'
                f' rowspan="{rowspan}">{fam}</td>'
            )
            html += (
                f'<td style="background:{fam_c};padding:8px 12px;text-align:center;"'
                f' rowspan="{rowspan}">{fam_count}</td>'
            )
            html += (
                f'<td style="background:{fam_c};padding:8px 12px;text-align:center;"'
                f' rowspan="{rowspan}">{fam_pct}%</td>'
            )
            first_row = False

        display_kw = kw if not kw_is_placeholder else "—"
        html += f'<td style="background:{kw_c};padding:8px 12px;">{display_kw}</td>'
        html += f'<td style="background:{kw_c};padding:8px 12px;text-align:center;">{kw_count}</td>'
        html += f'<td style="background:{kw_c};padding:8px 12px;text-align:center;">{kw_pct}%</td>'
        html += f'<td style="padding:8px 12px;">{pids_html}</td>'
        html += "</tr>\n"

        # ---- 同时存 raw data 给 Excel 下载用 ----
        table_rows.append({
            "_fam": fam, "_fam_n": fam_count, "_fam_p": f"{fam_pct}%",
            "_kw": kw if not kw_is_placeholder else "",
            "_kw_n": kw_count if not kw_is_placeholder else "",
            "_kw_p": f"{kw_pct}%" if not kw_is_placeholder else "",
            "涉及产品": re.sub(r"<[^>]+>", "", pids_html),
        })

html += "</tbody></table>"
st.markdown(html, unsafe_allow_html=True)

# ============================================================
# 下载按钮
# ============================================================
output = io.BytesIO()
out_wb = openpyxl.Workbook()

# Sheet 1: 分类详情
ws1 = out_wb.active
ws1.title = "分类详情"
ws1.append(["产品名", "大类", "细分", "Amazon链接", "方法"])
for r in results:
    m = "⚡curl_cffi" if r.get("方法")=="cffi" else "🐢Playwright" if r.get("方法")=="pw" else ""
    ws1.append([r["产品名"], r["大类"],
                r.get("细分_clean", ""), r.get("搜索URL", ""), m])

# Sheet 2: 层级统计
ws2 = out_wb.create_sheet("香型层级统计")
ws2.append(["大类", "大类数量", "大类占比", "细分", "细分数量", "细分占比", "涉及产品编号"])
for r in table_rows:
    pids = re.sub(r"<[^>]+>", "", r["涉及产品"])
    ws2.append([r["_fam"], r["_fam_n"], r["_fam_p"],
                r["_kw"], r["_kw_n"], r["_kw_p"], pids])

out_wb.save(output)
output.seek(0)

st.download_button(
    label="📥 下载统计 Excel",
    data=output,
    file_name="香型分类_统计结果.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)
