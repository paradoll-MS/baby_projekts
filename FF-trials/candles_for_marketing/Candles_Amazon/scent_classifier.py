#!/usr/bin/env python3
"""
香型分类脚本 v2
严格遵循 分类与搜索策略.md 的四阶段流水线：
  ① 搜索前排除（无香/电子蜡烛）→ 直接标记，不搜索
  ② 在线搜索（官网 → Amazon → 香水评测站 → Bing 通用搜索）
  ③ 文本清洗（去品牌名、wick材质、功能词等）
  ④ 关键词提取 → 香族映射（一级香族 + 二级香调，允许多标签）

输出：只保留 产品名、大类、细分 三列
  无香型/电子蜡烛只写在大类即可
"""

import json
import re
import sys
import time
import asyncio
from collections import defaultdict
from urllib.parse import quote
from pathlib import Path
import openpyxl
from bs4 import BeautifulSoup
from curl_cffi import requests as cffi_requests
from playwright.sync_api import sync_playwright

# 修复 Windows GBK 终端编码问题
if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

# ============================================================
# 0. 路径配置
# ============================================================
BASE_DIR = Path(__file__).parent
DICT_PATH = BASE_DIR / "fragrance_dict.json"
EXCLUSIONS_PATH = BASE_DIR / "brand_exclusions.json"
INPUT_EXCEL = BASE_DIR / "260722_久谦中台_跨境数据库_亚马逊_家居厨房_家居装饰_蜡烛_蜡烛_商品_销售额_年_backup.xlsx"
OUTPUT_EXCEL = BASE_DIR / "香型分类_30条_requests.xlsx"

# ============================================================
# 0. 加载配置
# ============================================================
with open(DICT_PATH, "r", encoding="utf-8") as f:
    fragrance_dict = json.load(f)

with open(EXCLUSIONS_PATH, "r", encoding="utf-8") as f:
    exclusions = json.load(f)

# 无香检测词（阶段①）
UNSCENTED_PATTERNS = (
    exclusions["unscented_detection"]["patterns"]
    + exclusions["unscented_detection"]["electronic"]
)

# 清洗模式（阶段③）
CLEAN_PATTERNS = exclusions["clean_patterns"]

# 香族映射
FAMILIES = fragrance_dict["families"]
MATCH_RULES = fragrance_dict["match_rules"]
WORD_BOUNDARY_WORDS = set(MATCH_RULES.get("word_boundary_for", []))

# 构建 关键词 → 香族名 的映射表
KEYWORD_TO_FAMILY = {}
# 提取中文简称（如 "花香 Floral" → "花香"）
FAMILY_CN = {}
for family_name, info in FAMILIES.items():
    cn = family_name.split()[0]  # "花香 Floral" → "花香"
    FAMILY_CN[family_name] = cn
    for kw in info["keywords"]:
        KEYWORD_TO_FAMILY[kw.lower()] = family_name

# 按长度降序排列所有关键词（长词优先匹配）
ALL_KEYWORDS_SORTED = sorted(KEYWORD_TO_FAMILY.keys(), key=len, reverse=True)

# ===== Playwright 浏览器（线程本地存储，避免 greenlet 跨线程报错） =====
import threading
_thread_local = threading.local()
_pw_instances = []  # 追踪所有实例，用于清理

def _get_browser():
    if not hasattr(_thread_local, 'browser') or _thread_local.browser is None:
        pw = sync_playwright().start()
        _thread_local.pw = pw
        _thread_local.browser = pw.chromium.launch(headless=True,
            args=["--disable-blink-features=AutomationControlled","--no-sandbox"])
        _pw_instances.append(pw)
    return _thread_local.browser

def cleanup_playwright():
    """关闭所有 Playwright 实例，释放资源。"""
    for pw in _pw_instances:
        try:
            pw.stop()
        except Exception:
            pass
    _pw_instances.clear()

def _new_page():
    ctx = _get_browser().new_context(
        user_agent="Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.0 Safari/605.1.15",
        viewport={"width": 800, "height": 600}, locale="en-US")
    # 只拦图片/字体/媒体（CSS 和 JS 不能拦——Akamai 验证需要 JS）
    ctx.route("**/*.{png,jpg,jpeg,gif,webp,svg,ico,woff,woff2,ttf,otf,eot,mp4,mp3,avi}",
              lambda r: r.abort())
    ctx.route("**/*analytics*", lambda r: r.abort())
    return ctx, ctx.new_page()


# ============================================================
# 阶段①：搜索前排除
# ============================================================
def detect_unscented(text: str) -> str | None:
    """
    检测是否为无香型/电子蜡烛。
    返回 None 表示有香，需要继续搜索；
    返回字符串表示无香类型标记。
    """
    text_lower = text.lower()

    # 先检测电子/LED
    for pat in exclusions["unscented_detection"]["electronic"]:
        pat_clean = pat.replace("-", r"[\-]?").replace(" ", r"\s+")
        if re.search(pat_clean, text_lower, re.IGNORECASE):
            return "无香型（电子/LED）"

    # 再检测无香
    for pat in exclusions["unscented_detection"]["patterns"]:
        pat_clean = pat.replace("-", r"[\-]?").replace(" ", r"\s+")
        if re.search(pat_clean, text_lower, re.IGNORECASE):
            return "无香型"

    return None  # 有香，需要搜索


# ============================================================
# 阶段②：在线搜索
# ============================================================
def extract_brand(product_name: str) -> str | None:
    """从商品名中尝试提取品牌名"""
    brands = CLEAN_PATTERNS["brands"]
    text_lower = product_name.lower()
    # 按品牌名长度降序匹配（长品牌名优先）
    sorted_brands = sorted(brands, key=len, reverse=True)
    for brand in sorted_brands:
        if brand.lower() in text_lower:
            return brand
    return None


def fetch_page_text(page) -> str:
    """从 Playwright 页面提取标题 + bullet points + Scent"""
    text_parts = []
    try:
        t = page.locator("#productTitle").first.inner_text(timeout=5000)
        text_parts.append(t.strip())
    except Exception:
        pass
    bp_count = 0
    skip = [r'^Made (with|in|from)', r'^(Quality|Consistent)',
            r'^(Long burning|Decorative)',
            r'^About (this|the) (brand|item)', r'^From the (brand|manufacturer)']
    try:
        for bp in page.locator("#feature-bullets li, #featurebullets_feature_div li").all():
            try:
                t = bp.inner_text(timeout=2000).strip()
            except Exception:
                continue
            if t and len(t) > 15 and not any(re.search(p, t, re.IGNORECASE) for p in skip):
                text_parts.append(t)
                bp_count += 1
                if bp_count >= 7:
                    break
    except Exception:
        pass
    ALL_KW = set(KEYWORD_TO_FAMILY.keys())
    try:
        for row in page.locator("tr").all()[:60]:
            try:
                th = row.locator("th").first; td = row.locator("td").first
                if th.count() and td.count():
                    label = th.inner_text(timeout=1000).strip().lower()
                    if label in ("scent","scent name","fragrance","fragrance name","aroma"):
                        value = td.inner_text(timeout=1000).strip().lower()
                        if value in ALL_KW:
                            text_parts.append(f"Scent: {value}")
                        break
            except Exception:
                continue
    except Exception:
        pass
    r = " ".join(text_parts)
    return re.sub(r'\s+', ' ', r).strip()[:2000] if r else ""


def _parse_amazon_search(html: str) -> list[str]:
    """从 Amazon 搜索页 HTML 提取 ASIN 列表"""
    soup = BeautifulSoup(html, "html.parser")
    asins = []
    for link in soup.select("a[href*=\"/dp/\"]"):
        href = link.get("href", "")
        m = re.search(r'/dp/([A-Z0-9]{10})', href)
        if m and m.group(1) not in asins and "aax-" not in href:
            asins.append(m.group(1))
        if len(asins) >= 8:
            break
    return asins


def _parse_product_html(html: str) -> str:
    """从 Amazon 商品页 HTML 提取文本"""
    soup = BeautifulSoup(html, "html.parser")
    text_parts = []
    title = soup.select_one("#productTitle")
    if title:
        text_parts.append(title.get_text(strip=True))
    bp_count = 0
    skip = [r'^Made (with|in|from)', r'^(Quality|Consistent)',
            r'^(Long burning|Decorative)',
            r'^About (this|the) (brand|item)', r'^From the (brand|manufacturer)']
    for bp in soup.select("#feature-bullets li, #featurebullets_feature_div li"):
        t = bp.get_text(strip=True)
        if t and len(t) > 15 and not any(re.search(p, t, re.IGNORECASE) for p in skip):
            text_parts.append(t)
            bp_count += 1
            if bp_count >= 7:
                break
    ALL_KW = set(KEYWORD_TO_FAMILY.keys())
    for row in soup.select("tr"):
        th = row.select_one("th"); td = row.select_one("td")
        if th and td and th.get_text(strip=True).lower() in \
           ("scent","scent name","fragrance","fragrance name","aroma"):
            v = td.get_text(strip=True).lower()
            if v in ALL_KW:
                text_parts.append(f"Scent: {v}")
            break
    r = " ".join(text_parts)
    return re.sub(r'\s+', ' ', r).strip()[:2000] if r else ""


def _search_keywords(product_name: str) -> tuple[list[str], set, str]:
    """生成搜索关键词"""
    words = [w.strip(",.()[]\"'!?:;/-") for w in product_name.split()]
    stop_words = {"for","with","and","the","a","an","in","on","of","to","up","by","is","or","as","at","from","&"}
    key_words = [w for w in words[:12] if w.lower() not in stop_words and len(w) > 1]
    origin_kw_set = set(w.lower() for w in key_words if len(w) > 3)
    terms = [" ".join(key_words[:8])]
    if len(key_words) > 5: terms.append(" ".join(key_words[:5]))
    if len(key_words) > 3: terms.append(" ".join(key_words[:3]))
    return terms, origin_kw_set, extract_brand(product_name) or ""


# ---- 方式一：curl_cffi（快，Akamai 不封时 ~3s/请求） ----
def search_amazon_cffi(product_name: str) -> tuple[str, str] | None:
    search_terms, origin_kw_set, brand = _search_keywords(product_name)
    for st in search_terms:
        try:
            r = cffi_requests.get(
                f"https://www.amazon.com/s?k={quote(st)}&i=garden",
                impersonate="safari17_0", timeout=12)
        except Exception:
            continue
        if r.status_code != 200 or len(r.text) < 10000:
            continue
        asins = _parse_amazon_search(r.text)
        if not asins:
            continue
        for asin in asins:
            try:
                r2 = cffi_requests.get(
                    f"https://www.amazon.com/dp/{asin}",
                    impersonate="safari17_0", timeout=12)
            except Exception:
                continue
            if r2.status_code != 200 or len(r2.text) < 10000:
                continue
            pt = _parse_product_html(r2.text)
            if pt and len(pt) >= 80:
                pt_lower = pt.split("|")[0].split(". ")[0].lower()
                if sum(1 for kw in origin_kw_set if kw in pt_lower) >= 1 \
                   or brand.lower() in pt_lower:
                    return pt, f"https://www.amazon.com/dp/{asin}"
        for asin in asins[:3]:
            try:
                r2 = cffi_requests.get(
                    f"https://www.amazon.com/dp/{asin}",
                    impersonate="safari17_0", timeout=12)
                if r2.status_code == 200 and len(r2.text) >= 10000:
                    pt = _parse_product_html(r2.text)
                    if pt and len(pt) >= 80:
                        return pt, f"https://www.amazon.com/dp/{asin}"
            except Exception:
                continue
    return None


# ---- 方式二：Playwright（慢但稳，~7s/请求，能过 Akamai） ----
def search_amazon_playwright(product_name: str) -> tuple[str, str] | None:
    search_terms, origin_kw_set, brand = _search_keywords(product_name)
    ctx, page = _new_page()
    try:
        for st in search_terms:
            try:
                page.goto(f"https://www.amazon.com/s?k={quote(st)}&i=garden",
                          wait_until="domcontentloaded", timeout=20000)
                page.wait_for_timeout(2000)
            except Exception:
                continue
            asins = []
            try:
                for link in page.locator("a[href*=\"/dp/\"]").all():
                    try:
                        href = link.get_attribute("href", timeout=1000) or ""
                    except Exception:
                        continue
                    m = re.search(r'/dp/([A-Z0-9]{10})', href)
                    if m and m.group(1) not in asins and "aax-" not in href:
                        asins.append(m.group(1))
                    if len(asins) >= 8:
                        break
            except Exception:
                pass
            for asin in asins:
                try:
                    page.goto(f"https://www.amazon.com/dp/{asin}",
                              wait_until="domcontentloaded", timeout=15000)
                    page.wait_for_timeout(1500)
                except Exception:
                    continue
                pt = fetch_page_text(page)
                if pt and len(pt) >= 80:
                    pt_lower = pt.split("|")[0].split(". ")[0].lower()
                    if sum(1 for kw in origin_kw_set if kw in pt_lower) >= 1 \
                       or brand.lower() in pt_lower:
                        return pt, f"https://www.amazon.com/dp/{asin}"
            for asin in asins[:3]:
                try:
                    page.goto(f"https://www.amazon.com/dp/{asin}",
                              wait_until="domcontentloaded", timeout=12000)
                    page.wait_for_timeout(1000)
                    pt = fetch_page_text(page)
                    if pt and len(pt) >= 80:
                        return pt, f"https://www.amazon.com/dp/{asin}"
                except Exception:
                    continue
    finally:
        page.close()
        ctx.close()
    return None


def search_amazon_direct(product_name: str) -> tuple[str, str, str] | None:
    """双接口：先试 curl_cffi（快），失败切 Playwright（稳）。
    返回 (页面文本, URL, 方法名)"""
    result = search_amazon_cffi(product_name)
    if result:
        return (result[0], result[1], "cffi")
    pw = search_amazon_playwright(product_name)
    if pw:
        return (pw[0], pw[1], "pw")
    return None


def search_product_online(product_name: str) -> tuple[str, str, str]:
    result = search_amazon_direct(product_name)
    if result:
        return f"{product_name} | {result[0]}", result[1], result[2]
    return product_name, "无搜索结果", "none"


# ============================================================
# 大类收敛 + Mix 检测
# ============================================================

def _resolve_classification(families: list[str], keywords: list[str],
                            search_only: set[str]) -> tuple[str, list[str]]:
    """大类收敛 + Mix 检测。
    返回 (primary_family, filtered_keywords)。

    规则：
    1. len(keywords) >= 8 → "混合 Mix"，保留全部关键词
    2. 否则按加权得分选最高大类：
       - 商品名自带关键词（不在 search_only 中）权重 2
       - 仅搜索来的关键词（在 search_only 中）权重 1
    3. 过滤：只保留属于选中大类的关键词
    """
    if len(keywords) >= 8:
        return ("混合 Mix", list(keywords))

    # 加权计分
    family_scores = defaultdict(int)
    for kw in keywords:
        fam = KEYWORD_TO_FAMILY.get(kw.lower())
        if fam:
            weight = 1 if kw.lower() in search_only else 2
            family_scores[fam] += weight

    if not family_scores:
        return ("未识别", [])

    # 最高分大类
    primary = max(family_scores, key=lambda f: (family_scores[f], f))

    # 过滤：只保留属于选中大类的关键词
    filtered = [kw for kw in keywords
                if KEYWORD_TO_FAMILY.get(kw.lower()) == primary]

    return (primary, filtered)


# ============================================================
# 异步批量分类（curl_cffi 并发，1-2s/条）
# ============================================================

def _classify_text(product_name: str, combined_text: str,
                   search_url: str = "", method: str = "") -> dict:
    """核心分类逻辑：阶段③清洗 + 阶段④关键词提取 → 香族映射。
    与 classify_product 共用同一套清洗/提取逻辑。"""
    # 阶段③
    cleaned = clean_text(combined_text)
    cleaned_name_only = clean_text(product_name)
    # 阶段④
    families, keywords, search_only = extract_fragrance_with_source(
        cleaned, cleaned_name_only)

    SKIP_IN_SECONDARY = {"fresh", "clean", "cotton", "cake"}
    keywords = [kw for kw in keywords if kw not in SKIP_IN_SECONDARY]
    search_only = {kw for kw in search_only if kw not in SKIP_IN_SECONDARY}

    # 大类收敛 + Mix 检测
    primary_family, filtered_keywords = _resolve_classification(
        families, keywords, search_only)

    result = {"产品名": product_name, "大类": primary_family,
              "细分": "", "搜索URL": search_url, "方法": method}

    if filtered_keywords:
        def fmt_kw(kw, is_search_only):
            fam = KEYWORD_TO_FAMILY.get(kw.lower(), "")
            cn = FAMILY_CN.get(fam, "")
            label = f"{kw}({cn})" if cn else kw
            return f"{label}[搜]" if is_search_only else label
        kw_parts = [fmt_kw(kw, kw.lower() in search_only) for kw in filtered_keywords]
        result["细分"] = ", ".join(kw_parts)
    return result


def _classify_one_sync(product_name: str) -> dict:
    """同步分类单个商品（curl_cffi 优先，失败自动切 Playwright）。"""
    unscented_type = detect_unscented(product_name)
    if unscented_type:
        return {"产品名": product_name, "大类": unscented_type,
                "细分": "", "搜索URL": "", "方法": "none"}

    # 一次调用，内部已有 cffi → pw fallback
    search_text, search_url, method = search_product_online(product_name)
    if search_text != product_name:
        return _classify_text(product_name, search_text, search_url, method)

    return {"产品名": product_name, "大类": "未识别", "细分": "",
            "搜索URL": "", "方法": "none"}


async def _classify_one_async(sem: asyncio.Semaphore, product_name: str,
                              start_delay: float = 0):
    """在线程池中运行同步分类，实现并发。"""
    if start_delay > 0:
        await asyncio.sleep(start_delay)

    async with sem:
        # 在默认线程池中执行同步函数
        return await asyncio.to_thread(_classify_one_sync, product_name)


async def batch_classify_products(products: list[str],
                                  max_concurrent: int = 4,
                                  progress_cb=None) -> list[dict]:
    """异步批量分类。用线程池并发跑同步 curl_cffi，错开启动避免限流。"""
    sem = asyncio.Semaphore(max_concurrent)
    stagger = 0.3  # 每件商品错开 0.3s 启动
    tasks = [
        _classify_one_async(sem, p, start_delay=i * stagger)
        for i, p in enumerate(products)
    ]
    results = await asyncio.gather(*tasks)
    return list(results)


# ============================================================
# 阶段③：文本清洗
# ============================================================
def clean_text(text: str) -> str:
    """
    从搜索文本中移除干扰项：品牌名、wick材质、蜡基、功能词、场景词、规格词。
    顺序很重要：先清理长的多词短语，再清理单词语。
    """
    text_lower = text.lower()

    # 把所有清洗目标合并，按长度降序排列
    all_clean_targets = []

    for category in ["brands", "wick_materials", "wax_bases",
                     "functional_words", "scene_words", "spec_words"]:
        for item in CLEAN_PATTERNS.get(category, []):
            all_clean_targets.append(item.lower())

    # 额外补充清洗目标（不在 JSON 中但需要处理的模式）
    extra_cleans = [
        # 蜡基变体：处理 "Coconut & Soy Blend Wax" 这类带 & 的写法
        "coconut & soy blend wax",
        "coconut and soy blend wax",
        "coconut soy blend wax",
        "coconut soy wax blend",
        "soy & coconut wax",
        "soy and coconut wax",
        "soy wax blend",
        "coconut wax blend",
        "coconut apricot wax",
        "coconut beeswax",
        # 装饰/派对蜡烛（非香调蜡烛，不含香味）
        "cake candle",
        "cake candles",
        "cake decoration",
        "cake decorations",
        "cake topper",
        "cake toppers",
        "birthday candle",
        "birthday candles",
        "celebration candle",
        "celebration candles",
        "party candle",
        "party candles",
        "decorative candle",
        "decorative candles",
        "decoration candle",
        "decoration candles",
    ]
    all_clean_targets.extend(extra_cleans)

    # 按长度降序（长短语优先替换）
    all_clean_targets.sort(key=len, reverse=True)

    for target in all_clean_targets:
        # 构建灵活匹配的正则：连字符可替换为 [ -]?，空格可多个，& 前后可能有空格
        pattern = (
            target
            .replace("-", r"[\-]?")
            .replace(" & ", r"\s*[&and]+\s*")
            .replace(" and ", r"\s*[&and]+\s*")
            .replace(" ", r"\s+")
        )
        # 使用词边界防止部分匹配
        text_lower = re.sub(r'\b' + pattern + r'\b', ' ', text_lower)

    # 清理多余空格
    text_lower = re.sub(r'\s+', ' ', text_lower).strip()

    return text_lower


# ============================================================
# 阶段④：关键词提取 → 香族映射
# ============================================================
def extract_fragrance_with_source(
    cleaned_text: str, product_name_cleaned: str
) -> tuple[list[str], list[str], set[str]]:
    """
    从清洗后的文本中提取香调关键词，映射到香族。
    同时区分：哪些关键词来自商品名本身，哪些仅来自搜索结果。

    返回 (一级香族列表, 二级香调列表, 仅来自搜索的关键词集合)。

    匹配规则：
      - 长词优先：fireside 先于 fir 被匹配
      - 短词词边界：fir/tea/salt/rose/pine/cake/spring 使用 \b 防止子串误判
      - 允许多标签
    """
    text = cleaned_text.lower()
    matched_keywords = set()
    matched_families = set()

    # 记录已匹配位置，防止同一位置重复匹配
    matched_spans = []

    for keyword in ALL_KEYWORDS_SORTED:
        kw_lower = keyword.lower()

        if kw_lower in WORD_BOUNDARY_WORDS:
            pattern = r'\b' + re.escape(kw_lower) + r'\b'
        else:
            pattern = re.escape(kw_lower)

        for match in re.finditer(pattern, text, re.IGNORECASE):
            start, end = match.start(), match.end()

            overlaps = any(
                not (end <= s or start >= e) for s, e in matched_spans
            )

            if not overlaps:
                matched_keywords.add(kw_lower)
                family = KEYWORD_TO_FAMILY[kw_lower]
                if family:
                    matched_families.add(family)
                matched_spans.append((start, end))

    # === 区分来源：哪些关键词仅来自搜索结果 ===
    search_only_keywords = set()
    for kw in matched_keywords:
        kw_lower = kw.lower()
        if kw_lower in WORD_BOUNDARY_WORDS:
            pat = r'\b' + re.escape(kw_lower) + r'\b'
        else:
            pat = re.escape(kw_lower)
        # 关键词不在商品名中 → 仅来自搜索
        if not re.search(pat, product_name_cleaned, re.IGNORECASE):
            search_only_keywords.add(kw_lower)

    family_list = sorted(matched_families)
    keyword_list = sorted(matched_keywords)

    return family_list, keyword_list, search_only_keywords


# ============================================================
# 主流水线
# ============================================================
def classify_product(product_name: str, verbose: bool = True) -> dict:
    """
    对单个商品执行完整的四阶段分类流水线。

    返回 {
        "产品名": str,
        "大类": str,       # 一级香族 或 "无香型" / "无香型（电子/LED）"
        "细分": str,       # 二级香调关键词
        "无香判定": str,   # "是" / "否"
        "来源": str,       # "关键词匹配" / "在线搜索" / "直接排除"
        "搜索URL": str,    # 在线搜索时的来源URL
    }
    """
    # ========== 阶段①：搜索前排除 ==========
    unscented_type = detect_unscented(product_name)
    if unscented_type:
        if verbose:
            print(f"  → {unscented_type}（跳过搜索）")
        return {
            "产品名": product_name, "大类": unscented_type,
            "细分": "", "搜索URL": "", "方法": "none",
        }

    # ========== 阶段②：在线搜索 ==========
    if verbose:
        print("  → 正在在线搜索...")
    search_text, search_url, method = search_product_online(product_name)

    result = {
        "产品名": product_name, "大类": "", "细分": "",
        "搜索URL": search_url, "方法": method,
    }

    # 合并搜索文本 + 原始商品名（商品名中也可能包含香调信息）
    combined_text = f"{product_name} | {search_text}" if search_text else product_name

    if verbose:
        text_preview = combined_text[:120].replace("\n", " ")
        print(f"  → 搜索文本: {text_preview}...")

    # ========== 阶段③：文本清洗 ==========
    cleaned = clean_text(combined_text)
    cleaned_name_only = clean_text(product_name)

    # ========== 阶段④：关键词提取 → 香族映射 ==========
    families, keywords, search_only = extract_fragrance_with_source(
        cleaned, cleaned_name_only
    )

    # 剔除不适合出现在二级香调中的通用词（一级香族保留）
    SKIP_IN_SECONDARY = {"fresh", "clean", "cotton", "cake"}
    keywords = [kw for kw in keywords if kw not in SKIP_IN_SECONDARY]
    search_only = {kw for kw in search_only if kw not in SKIP_IN_SECONDARY}

    # 大类收敛 + Mix 检测
    primary_family, filtered_keywords = _resolve_classification(
        families, keywords, search_only)

    result["大类"] = primary_family
    result["搜索URL"] = search_url

    if filtered_keywords:
        def fmt_kw(kw, is_search_only):
            fam = KEYWORD_TO_FAMILY.get(kw.lower(), "")
            cn = FAMILY_CN.get(fam, "")
            label = f"{kw}({cn})" if cn else kw
            return f"{label}[搜]" if is_search_only else label
        kw_parts = [fmt_kw(kw, kw.lower() in search_only) for kw in filtered_keywords]
        result["细分"] = ", ".join(kw_parts)
    else:
        result["细分"] = ""
        result["搜索URL"] = search_url

    if verbose:
        print(f"  → 结果: 大类={result['大类']} | 细分={result['细分']}")

    return result


# ============================================================
# 批量处理 + 导出
# ============================================================
def main():
    print("=" * 60)
    print("香型分类脚本 v2")
    print("=" * 60)

    # 读取原始 Excel
    print(f"\n读取数据: {INPUT_EXCEL.name}")
    wb = openpyxl.load_workbook(INPUT_EXCEL)
    ws = wb.active

    # 商品名在第 9 列（0-indexed: 8），数据从第 3 行开始
    products = []
    for row_idx in range(3, ws.max_row + 1):
        product_name = ws.cell(row=row_idx, column=9).value
        if product_name and str(product_name).strip():
            products.append(str(product_name).strip())

    print(f"共 {len(products)} 条商品记录")
    print(f"本次处理前 30 条\n")

    N = min(30, len(products))
    results = []
    for i in range(N):
        product_name = products[i]
        try:
            result = classify_product(product_name, verbose=False)
        except Exception as e:
            result = {"产品名": product_name, "大类": "处理失败",
                      "细分": str(e)[:50], "搜索URL": ""}
        results.append(result)
        m = result.get("方法", "")
        tag = "⚡" if m == "cffi" else "🐢" if m == "pw" else ""
        print(f"[{i+1}/{N}] {tag} {result['大类'][:35]} | {product_name[:60]}")
        time.sleep(0.5)

    # 导出到新 Excel
    print(f"\n导出结果到: {OUTPUT_EXCEL.name}")

    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "香型分类结果"

    # 五列：产品名、大类、细分、Amazon链接、方法
    headers = ["产品名", "大类", "细分", "Amazon链接", "方法"]
    out_ws.append(headers)

    # 写入数据
    for i, r in enumerate(results):
        clean_sub = re.sub(r'\[搜\]', '', r["细分"]) if r["细分"] else ""
        method_label = "⚡curl_cffi" if r.get("方法") == "cffi" else \
                       "🐢Playwright" if r.get("方法") == "pw" else \
                       "—" if r.get("方法") == "none" else ""
        row_data = [r["产品名"], r["大类"], clean_sub, r["搜索URL"], method_label]
        out_ws.append(row_data)

    out_ws.column_dimensions["A"].width = 80
    out_ws.column_dimensions["B"].width = 30
    out_ws.column_dimensions["C"].width = 50
    out_ws.column_dimensions["D"].width = 50
    out_ws.column_dimensions["E"].width = 14

    # 冻结表头
    out_ws.freeze_panes = "A2"

    out_wb.save(OUTPUT_EXCEL)

    # 统计
    unscented_count = sum(1 for r in results if "无香" in r["大类"])
    classified_count = sum(1 for r in results
                           if r["大类"] and "无香" not in r["大类"]
                           and "未识别" not in r["大类"] and "失败" not in r["大类"])
    unknown_count = sum(1 for r in results if "未识别" in r["大类"])
    failed_count = sum(1 for r in results if "失败" in r["大类"])

    print(f"\n{'='*60}")
    print(f"处理完成！")
    print(f"  总计: {len(results)} 条")
    print(f"  无香型/电子: {unscented_count} 条")
    print(f"  已分类: {classified_count} 条")
    print(f"  未识别: {unknown_count} 条")
    print(f"  处理失败: {failed_count} 条")
    print(f"  结果文件: {OUTPUT_EXCEL}")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
